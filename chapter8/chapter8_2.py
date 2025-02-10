import pathlib
from openpyxl import load_workbook
from itertools import islice

my_path = pathlib.Path(r"chapter8\data")
data = tuple()
for p in my_path.glob("*.xlsx"):
    wb = load_workbook(p)
    ws = wb.active
    values = islice(ws.values, 1, None)
    data += tuple(values)

filename = "chapter8\売上集計.xlsx"
wb_summary = load_workbook(filename)
ws_data = wb_summary["data"]

for row in data:
    ws_data.append(row)
print(data)
wb_summary.save(filename)