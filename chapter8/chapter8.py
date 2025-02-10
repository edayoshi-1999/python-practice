# from openpyxl import Workbook
from openpyxl import load_workbook

# wb = Workbook()
# wb.save(r"chapter8\test.xlsx")
# print(type(wb))

# filename = r"chapter8\売上集計.xlsx"
# wb_summary = load_workbook(filename)
# print(wb_summary.sheetnames)
# print(wb_summary.active.title)
# print(wb_summary["summary"].title)
# for sheet in wb_summary:
#     print(sheet)

# ws = wb_summary.active
# c = ws["A1"]
# print(c, type(c))
# rng = ws["A1:B2"]
# print(rng, type(rng))

# c = ws.cell(2, 3)
# print(c.value)
# d = ws.cell(12, 6, value=100)
# print(d.value)
# wb_summary.save(filename)

# filename = r"chapter8\売上集計.xlsx"
# wb = load_workbook(filename)
# ws = wb.active
# for row in ws.rows:
#     print(row)
# for row in ws.iter_rows(min_row=1, max_row=2, min_col=3, max_col=4):
#     print(row)
# for col in ws.columns:
#     print(col)
# for col in ws.iter_cols(min_row=1, max_row=2, min_col=3, max_col=4):
#     print(col)
# for row in ws.values:
#     print(row)

# filename = r"chapter8\売上集計.xlsx"
# wb = load_workbook(filename)
# ws = wb.active
# c = ws["C2"]
# print(c.value)
# print(c.row)
# print(c.column)
# record = ["2020/03/10", "店舗A", "商品１", 66, 66000]
# ws.append(record)
# c = ws["C13"]
# c.value = "商品4"
# wb.save(filename)